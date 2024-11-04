


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from bs4 import BeautifulSoup
import time
import os
from openpyxl import Workbook
import re



def wait_for_user_action(driver, message):
    print(message)
    input("Nhấn Enter sau khi bạn đã hoàn thành hành động...")
    time.sleep(5)  # Đợi thêm 5 giây sau khi người dùng nhấn Enter


def get_next_idea_number():
    if not os.path.exists('idea_links.txt'):
        return 1
    with open('idea_links.txt', 'r', encoding='utf-8') as f:
        lines = f.readlines()
        if not lines:
            return 1
        for line in reversed(lines):
            parts = line.strip().split(':')
            if len(parts) >= 2:
                try:
                    number = int(''.join(filter(str.isdigit, parts[0])))
                    return number + 1
                except ValueError:
                    continue
    return 1  # Nếu không tìm thấy số hợp lệ, bắt đầu từ 1

def is_link_exists(url):
    if not os.path.exists('idea_links.txt'):
        return False
    with open('idea_links.txt', 'r', encoding='utf-8') as f:
        return any(url in line for line in f)
    
def extract_page_info(driver):
    title = driver.title
    description = driver.find_element(By.XPATH, '//meta[@name="description"]').get_attribute('content')
    return title, description

from openpyxl import Workbook, load_workbook

def save_to_excel(ideas):
    file_name = "ideas.xlsx"
    if os.path.exists(file_name):
        # Nếu file đã tồn tại, mở nó và xác định hàng bắt đầu ghi mới
        wb = load_workbook(file_name)
        ws = wb.active
        start_row = ws.max_row + 1
    else:
        # Nếu file chưa tồn tại, tạo mới với các tiêu đề
        wb = Workbook()
        ws = wb.active
        ws.title = "Ideas"
        headers = ["STT", "Link", "Title", "Meta Description", "Description", "Built with", "About"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        start_row = 2

    # Ghi dữ liệu mới
    for row, idea in enumerate(ideas, start=start_row):
        ws.cell(row=row, column=1, value=idea['stt'])
        ws.cell(row=row, column=2, value=idea['link'])
        ws.cell(row=row, column=3, value=idea['title'])
        ws.cell(row=row, column=4, value=idea['meta_description'])
        ws.cell(row=row, column=5, value=idea['description'])
        ws.cell(row=row, column=6, value=idea['built_with'])
        ws.cell(row=row, column=7, value=idea['about'])

    wb.save(file_name)
    print(f"Đã lưu dữ liệu vào file ideas.xlsx (từ hàng {start_row})")

# ... rest of the code remains the same ...
# Khởi tạo danh sách để lưu các ý tưởng
ideas = []

# B1: Truy cập trang này: https://ai.google.dev/competition/
url = "https://ai.google.dev/competition/"
driver = webdriver.Chrome()
driver.get(url)

# B2: Đợi và kích vào "Vote for your favorite apps"
# B2: Đợi vài giây, load xong trang thì Kích vào element: 
# <a class="gemini-btn gemini-gradient  gemini-type-btn gc-analytics-event" data-category="content_click" data-label="https://ai.google.dev/competition/vote" data-action="Vote for your favorite apps" href="/competition/vote" rel="">
#       <span>Vote for your favorite apps</span>
#     </a>
# XPath: //*[@id="gemini-overview"]/section/div[2]/div/div/article/div[4]/a

try:
    vote_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//span[text()='Vote for your favorite apps']"))
    )
    driver.execute_script("arguments[0].click();", vote_button)
    time.sleep(5)  # Đợi trang mới tải xong
except Exception as e:
    print(f"Không thể nhấp vào 'Vote for your favorite apps': {e}")
    driver.save_screenshot("error_vote_apps.png")
    driver.quit()
    exit()

# B3: Đợi vài giây, load xong trang thì Kích vào element: 
# <div class="gemini-btn gemini-secondary gemini-start-button gemini-type-btn">
#       <span>Start&nbsp;voting</span>
#     </div>
try:
    start_voting = WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.XPATH, "//div[@class='gemini-btn gemini-secondary gemini-start-button gemini-type-btn']"))
    )
    driver.execute_script("arguments[0].click();", start_voting)
    print("Đã nhấp vào 'Start voting' bằng XPath đầy đủ")
    time.sleep(5)
except Exception as e:
    print(f"Không thể nhấp vào 'Start voting': {e}")
    wait_for_user_action(driver, "Vui lòng nhấp vào 'Start voting' thủ công.")

# B4 and B5: Collect ideas and navigate back
# B4:Đợi vài giây Ấn vào element: <div class="gemini-vote-card-glow"></div> 
# Sau khi ấn thì Viết vào file text. 
# - Ý tưởng thứ mấy: 
# - lưu lại đường link của trang 
# - Quét lưu VĂN BẢN (KO PHẢI HTML)(kể cả đường đường link youtube ) ở XPath: 
# //*[@id="gc-wrapper"]/main/devsite-content/article/div[2]/devsite-gemini-project-page/div[2]
# Note CHỈNH SỬA THÊM: 

# B5:
# - Ấn nút mũi tên quay lại 
# - Sau đó lặp lại bước 4 số lần là: 3 
for i in range(2):
    # Lấy số thứ tự mới cho ý tưởng
    idea_number = get_next_idea_number()
    try:
        print(f"Đang tìm ý tưởng thứ {i+1}...")
        element = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "div.gemini-vote-card-glow"))
        )
        print(f"Đã tìm thấy ý tưởng thứ {i+1}. Đang scroll đến element...")
        driver.execute_script("arguments[0].scrollIntoView(true);", element)
        time.sleep(2)
        print("Đang nhấp vào element...")
        ActionChains(driver).move_to_element(element).click().perform()
        print("Đã nhấp vào element. Đang đợi 5 giây...")
        time.sleep(5)
        
        # Lưu đường link của trang
        current_url = driver.current_url
        print(current_url)
        
        # Kiểm tra xem link đã tồn tại chưa
        if is_link_exists(current_url):
            print(f"Link {current_url} đã tồn tại. Bỏ qua.")
            driver.back()
            time.sleep(5)
            continue
        
        # Lưu link vào file idea_links.txt
        with open('idea_links.txt', 'a', encoding='utf-8') as f:
            f.write(f"{idea_number}. {current_url}\n")
        
        # Trích xuất title và description
        title, meta_description = extract_page_info(driver)

        
        print("Đang quét thông tin văn bản...")
        # Quét thông tin văn bản từ XPath
        text_element = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="gc-wrapper"]/main/devsite-content/article/div[2]/devsite-gemini-project-page/div[2]'))
        )
        
        # Sử dụng BeautifulSoup để trích xuất văn bản và link
        soup = BeautifulSoup(text_element.get_attribute('outerHTML'), 'html.parser')


        # 1. Trong phần XPath://*[@id="gc-wrapper"]/main/devsite-content/article/div[2]/devsite-gemini-project-page/div[2]/section[3]/div[1]/div[1]/p[2] này có các đoạn thẻ <br> thể hiện việc xuống dòng. 
        # - Trong output tôi lại thấy nó viết liền => tôi muốn nó cũng xuống dòng
        # - Bạn muốn giữ nguyên cấu trúc xuống dòng của các thẻ <br> trong output, mà không hiển thị thẻ HTML.

        
        def extract_description(div):
            if not div:
                return "Description not available"
            
            description = ""
            
            # Trích xuất tiêu đề "What it does"
            title = div.find('p', class_='gemini-description-label')
            if title:
                description += title.get_text(strip=True) + "\n\n"
            
            # Trích xuất nội dung mô tả, xử lý các thẻ <br>
            content = div.find('p', class_='gemini-type-b1')
            if content:
                for element in content.contents:
                    if element.name == 'br':
                        description += '\n'
                    elif isinstance(element, str):
                        description += element.strip()
                    else:
                        description += element.get_text(strip=True)
            
            return description.strip()

        # Sử dụng hàm này trong vòng lặp chính
        description_div = soup.find('div', class_='gemini-project-description')
        description = extract_description(description_div)

        def extract_tools(div):
            if not div:
                return "Tools information not available"
            tools = div.find_all('li', class_='gemini-gradient-text')
            return ", ".join([tool.get_text(strip=True) for tool in tools])

        def extract_about(div):
            if not div:
                return "About information not available"
            
            about_info = []
            
            # Trích xuất tiêu đề "Team"
            team_label = div.find('p', class_='gemini-description-label')
            if team_label:
                about_info.append(f"Team: {team_label.get_text(strip=True)}")
            
            # Trích xuất thông tin chi tiết
            details = div.find_all('div', class_='gemini-project-detail')
            for detail in details:
                label = detail.find('p', class_='gemini-type-t2')
                value = detail.find_all('p', class_='gemini-type-t2')[-1]  # Lấy phần tử cuối cùng
                if label and value:
                    about_info.append(f"{label.get_text(strip=True)}: {value.get_text(strip=True)}")
            
            return ", ".join(about_info) if about_info else "Detailed about information not available"

        # Trích xuất thông tin
        description_div = soup.find('div', class_='gemini-project-description')
        tools_div = soup.find('div', class_='gemini-project-tools')
        about_div = soup.find('div', class_='gemini-project-about')

        description = extract_description(description_div)
        tools = extract_tools(tools_div)
        about = extract_about(about_div)

        # Tổng hợp nội dung
        idea_text = f"- TITLE: {title}\n- ABSTRACT: {meta_description}\n\n*** Description:\n{description}\n\n*** Built with:\n{tools}\n\n*** About:\n{about}"

        print(idea_text)
        print(idea_text)
        # Viết vào file text
        with open('ideas.txt', 'a', encoding='utf-8') as f:
            f.write(f"Ý tưởng thứ {idea_number}:\n")  # Thay đổi ở đây
            f.write(f"Link: {current_url}\n")
            f.write(f"{idea_text}\n")
            f.write("\n" + "-"*50 + "\n\n")

        print(f"Đã thu thập và lưu ý tưởng thứ {idea_number}")  # Thay đổi ở đây

        # Cập nhật dictionary idea
        idea = {
            'stt': idea_number,
            'link': current_url,
            'title': title,
            'meta_description': meta_description,
            'description': description,
            'built_with': tools,
            'about': about
        }
        ideas.append(idea)
        
        # Quay lại trang trước
        driver.back()
        time.sleep(5)  # Tăng thời gian chờ sau khi quay lại
        
    except Exception as e:
        print(f"Lỗi khi thu thập ý tưởng thứ {idea_number}: {e}")  # Thay đổi ở đây
        driver.save_screenshot(f"error_idea_{idea_number}.png")  # Thay đổi ở đây
        wait_for_user_action(driver, f"Vui lòng xử lý vấn đề và nhấp vào ý tưởng thứ {idea_number} thủ công.")  # Thay đổi ở đây
# Sau khi kết thúc vòng lặp, gọi hàm lưu vào Excel
save_to_excel(ideas)

driver.quit()

# Kiểm tra nội dung file
if os.path.exists('ideas.txt'):
    with open('ideas.txt', 'r', encoding='utf-8') as f:
        content = f.read()
        if content.strip():
            print("File ideas.txt đã được tạo và có nội dung.")
        else:
            print("File ideas.txt tồn tại nhưng trống.")
else:
    print("File ideas.txt không tồn tại.")

print("Quá trình crawl đã hoàn tất.")