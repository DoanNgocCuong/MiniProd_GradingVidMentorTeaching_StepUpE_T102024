from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from bs4 import BeautifulSoup
import time
import os
from openpyxl import Workbook, load_workbook
import re
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException

def wait_for_user_action(driver, message):
    print(message)
    input("Nhấn Enter sau khi bạn đã hoàn thành hành động...")
    time.sleep(5)

def get_next_idea_number():
    if not os.path.exists('idea_links.txt'):
        return 1
    with open('idea_links.txt', 'r', encoding='utf-8') as f:
        lines = f.readlines()
        if not lines:
            return 1
        for line in reversed(lines):
            parts = line.strip().split('.')
            if len(parts) >= 2:
                try:
                    number = int(parts[0])
                    return number + 1
                except ValueError:
                    continue
    return 1

def is_link_exists(url):
    if not os.path.exists('idea_links.txt'):
        return False
    with open('idea_links.txt', 'r', encoding='utf-8') as f:
        return any(url in line for line in f)

def extract_page_info(driver):
    title = driver.title
    description = driver.find_element(By.XPATH, '//meta[@name="description"]').get_attribute('content')
    return title, description

def save_to_excel(ideas):
    file_name = "ideas.xlsx"
    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
        start_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Ideas"
        headers = ["STT", "Link", "Title", "Meta Description", "Description", "Built with", "About", "YouTube Links"]  # Thêm cột YouTube Links
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        start_row = 2

    for row, idea in enumerate(ideas, start=start_row):
        ws.cell(row=row, column=1, value=idea['stt'])
        ws.cell(row=row, column=2, value=idea['link'])
        ws.cell(row=row, column=3, value=idea['title'])
        ws.cell(row=row, column=4, value=idea['meta_description'])
        ws.cell(row=row, column=5, value=idea['description'])
        ws.cell(row=row, column=6, value=idea['built_with'])
        ws.cell(row=row, column=7, value=idea['about'])
        ws.cell(row=row, column=8, value=", ".join(idea['youtube_links']))  # Ghi các link YouTube vào cột

    wb.save(file_name)
    print(f"Đã lưu dữ liệu vào file ideas.xlsx (từ hàng {start_row})")

def extract_description(div):
    if not div:
        return "Description not available"
    
    description = ""
    
    title = div.find('p', class_='gemini-description-label')
    if title:
        description += title.get_text(strip=True) + "\n\n"
    
    content = div.find('p', class_='gemini-type-b1')
    if content:
        for element in content.contents:
            if element.name == 'br':
                description += '\n'
            elif isinstance(element, str):
                description += element.strip()
            else:
                description += element.get_text(strip=True)
    
    return description.strip()

def extract_tools(div):
    if not div:
        return "Tools information not available"
    tools = div.find_all('li', class_='gemini-gradient-text')
    return ", ".join([tool.get_text(strip=True) for tool in tools])

def extract_about(div):
    if not div:
        return "About information not available"
    
    about_info = []
    
    # Trích xuất tiêu đề "Team"
    team_label = div.find('p', class_='gemini-description-label')
    if team_label:
        about_info.append(f"Team: {team_label.get_text(strip=True)}")
    
    # Trích xuất thông tin chi tiết
    details = div.find_all('div', class_='gemini-project-detail')
    for detail in details:
        label = detail.find('p', class_='gemini-type-t2')
        value = detail.find_all('p', class_='gemini-type-t2')[-1]  # Lấy phần tử cuối cùng
        if label and value:
            about_info.append(f"{label.get_text(strip=True)}: {value.get_text(strip=True)}")
    
    return ", ".join(about_info) if about_info else "Detailed about information not available"
# ... rest of the code remains the same ...

def save_state(current_index):
    with open('crawl_state.txt', 'w') as f:
        f.write(str(current_index))

def load_state():
    if os.path.exists('crawl_state.txt'):
        with open('crawl_state.txt', 'r') as f:
            return int(f.read())
    return 0


def extract_youtube_links(soup):
    youtube_links = []
    
    # Tìm tất cả các thẻ devsite-video
    devsite_videos = soup.find_all('devsite-video')
    for video in devsite_videos:
        # Lấy video-id từ thuộc tính của devsite-video
        video_id = video.get('video-id')
        if video_id:
            youtube_links.append(f'https://www.youtube.com/watch?v={video_id}')
        
        # Tìm iframe bên trong devsite-video
        iframe = video.find('iframe')
        if iframe:
            src = iframe.get('src')
            if src and 'youtube.com/embed/' in src:
                video_id = src.split('/')[-1].split('?')[0]
                youtube_links.append(f'https://www.youtube.com/watch?v={video_id}')
    
    # Tìm tất cả các iframe trực tiếp (để đảm bảo)
    iframes = soup.find_all('iframe', src=lambda x: x and 'youtube.com/embed/' in x)
    for iframe in iframes:
        src = iframe.get('src')
        video_id = src.split('/')[-1].split('?')[0]
        youtube_link = f'https://www.youtube.com/watch?v={video_id}'
        if youtube_link not in youtube_links:
            youtube_links.append(youtube_link)
    
    return youtube_links


def crawl_ideas(driver, start_index, end_index):
    ideas = []
    actual_end_index = start_index
    
    for i in range(start_index, end_index):
        idea_number = get_next_idea_number()
        max_retries = 1
        for attempt in range(max_retries):
            try:
                print(f"Đang tìm ý tưởng thứ {idea_number}...")
                element = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "div.gemini-vote-card-glow"))
                )
                print(f"Đã tìm thấy ý tưởng thứ {idea_number}. Đang scroll đến element...")
                driver.execute_script("arguments[0].scrollIntoView(true);", element)
                time.sleep(2)
                print("Đang nhấp vào element...")
                ActionChains(driver).move_to_element(element).click().perform()
                print("Đã nhấp vào element. Đang đợi 5 giây...")
                time.sleep(5)
                
                current_url = driver.current_url
                print(current_url)
                
                if is_link_exists(current_url):
                    print(f"Link {current_url} đã tồn tại. Bỏ qua.")
                    driver.back()
                    time.sleep(5)
                    continue  # Skip this iteration

                with open('idea_links.txt', 'a', encoding='utf-8') as f:
                    f.write(f"{idea_number}. {current_url}\n")
                
                title, meta_description = extract_page_info(driver)
                
                print("Đang quét thông tin văn bản...")
                text_element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="gc-wrapper"]/main/devsite-content/article/div[2]/devsite-gemini-project-page/div[2]'))
                )
                
                soup = BeautifulSoup(text_element.get_attribute('outerHTML'), 'html.parser')

                description_div = soup.find('div', class_='gemini-project-description')
                tools_div = soup.find('div', class_='gemini-project-tools')
                about_div = soup.find('div', class_='gemini-project-about')

                description = extract_description(description_div)
                tools = extract_tools(tools_div)
                about = extract_about(about_div)

                # Trong vòng lặp chính của bạn
                youtube_links = extract_youtube_links(soup)  # Lấy các link YouTube
                print(f"Found YouTube links: {youtube_links}")
                
                idea_text = f"- TITLE: {title}\n- ABSTRACT: {meta_description}\n\n*** Description:\n{description}\n\n*** Built with:\n{tools}\n\n*** About:\n{about}"

                with open('ideas.txt', 'a', encoding='utf-8') as f:
                    f.write(f"Ý tưởng thứ {idea_number}:\n")
                    f.write(f"Link: {current_url}\n")
                    f.write(f"{idea_text}\n")
                    if youtube_links:  # Kiểm tra nếu có link YouTube
                        f.write("Link YouTube:\n")
                        for link in youtube_links:
                            f.write(f"{link}\n")
                    f.write("\n" + "-"*50 + "\n\n")

                print(f"Đã thu thập và lưu ý tưởng thứ {idea_number}")

                idea = {
                    'stt': idea_number,
                    'link': current_url,
                    'title': title,
                    'meta_description': meta_description,
                    'description': description,
                    'built_with': tools,
                    'about': about,
                    'youtube_links': youtube_links  # Thêm link YouTube vào dictionary
                }
                ideas.append(idea)
                
                driver.back()
                time.sleep(5)
                break  # Thêm break ở đây để thoát khỏi vòng lặp retry nếu thành công
                
                
            except (TimeoutException, NoSuchElementException, StaleElementReferenceException) as e:
                print(f"Lỗi khi thu thập ý tưởng thứ {idea_number} (Lần thử {attempt + 1}/{max_retries}): {e}")
                driver.save_screenshot(f"error_idea_{idea_number}_attempt_{attempt + 1}.png")
                if attempt == max_retries - 1:
                    print("Đã hết số lần thử. Đang quay lại trang chủ và thử lại từ đầu...")
                    try:
                        driver.get("https://ai.google.dev/competition/")
                        time.sleep(5)
                        
                        vote_button = WebDriverWait(driver, 20).until(
                            EC.element_to_be_clickable((By.XPATH, "//span[text()='Vote for your favorite apps']"))
                        )
                        driver.execute_script("arguments[0].click();", vote_button)
                        time.sleep(5)
                        
                        start_voting = WebDriverWait(driver, 20).until(
                            EC.element_to_be_clickable((By.XPATH, "//div[@class='gemini-btn gemini-secondary gemini-start-button gemini-type-t1-medium']"))
                        )
                        driver.execute_script("arguments[0].click();", start_voting)
                        print("Đã nhấp vào 'Start voting' bằng XPath đầy đủ")
                        time.sleep(5)
                        # Reset attempt counter to try again
                        attempt = -1
                        continue
                    except Exception as e:
                        print(f"Lỗi khi quay lại trang chủ: {e}")
                        wait_for_user_action(driver, "Vui lòng kiểm tra và điều hướng đến trang chính xác.")
                else:
                    time.sleep(5)

        if attempt == max_retries - 1:
            print(f"Không thể thu thập ý tưởng thứ {idea_number} sau khi thử lại. Tiếp tục với ý tưởng tiếp theo.")
            continue
        
        actual_end_index = i + 1
    
    return ideas, actual_end_index
def main():
    BATCH_SIZE = 20 # 3
    TOTAL_IDEAS = 3046
    start_index = load_state()

    driver = webdriver.Chrome()
    try:
        url = "https://ai.google.dev/competition/"
        driver.get(url)

        try:
            vote_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//span[text()='Vote for your favorite apps']"))
            )
            driver.execute_script("arguments[0].click();", vote_button)
            time.sleep(5)
        except Exception as e:
            print(f"Không thể nhấp vào 'Vote for your favorite apps': {e}")
            driver.save_screenshot("error_vote_apps.png")
            return

        try:
            start_voting = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//div[@class='gemini-btn gemini-secondary gemini-start-button gemini-type-t1-medium']"))
            )
            driver.execute_script("arguments[0].click();", start_voting)
            print("Đã nhấp vào 'Start voting' bằng XPath đầy đủ")
            time.sleep(5)
        except Exception as e:
            print(f"Không thể nhấp vào 'Start voting': {e}")
            wait_for_user_action(driver, "Vui lòng nhấp vào 'Start voting' thủ công.")

        while start_index < TOTAL_IDEAS:
            end_index = min(start_index + BATCH_SIZE, TOTAL_IDEAS)
            print(f"Crawling ideas from {start_index} to {end_index}")
            
            ideas, actual_end_index = crawl_ideas(driver, start_index, end_index)
            try:
                save_to_excel(ideas)
                save_state(actual_end_index)  # Use actual_end_index instead of end_index
                print(f"Completed batch {start_index} to {actual_end_index}")
                print(f"Current state: {actual_end_index}")
                print(f"Total ideas collected so far: {get_next_idea_number() - 1}")
            except PermissionError:
                print(f"Không thể lưu vào file Excel. Vui lòng đóng file nếu đang mở.")
                user_input = input("Nhấn Enter để thử lại hoặc 'q' để thoát: ")
                if user_input.lower() == 'q':
                    break
                continue  # Thử lại batch hiện tại
            
            start_index = actual_end_index
            
            time.sleep(5)

        print("Quá trình crawl đã hoàn tất.")
    finally:
        driver.quit()

if __name__ == "__main__":
    main()